import openpyxl
import time
import datetime

from base_func import get_data


def save_data(statistics_list):
    statistics_table = openpyxl.load_workbook(filename="statistics.xlsx")
    sheet = statistics_table['info']
    new_row = sheet.max_row + 1
    sheet.cell(row=new_row, column=1).value = statistics_list[0]
    sheet.cell(row=new_row, column=2).value = statistics_list[1]
    sheet.cell(row=new_row, column=3).value = statistics_list[2]
    sheet.cell(row=new_row, column=4).value = statistics_list[3]
    sheet.cell(row=new_row, column=5).value = statistics_list[4]
    sheet.cell(row=new_row, column=6).value = statistics_list[5]
    sheet.cell(row=new_row, column=7).value = statistics_list[6]
    sheet.cell(row=new_row, column=8).value = statistics_list[7]
    sheet.cell(row=new_row, column=9).value = statistics_list[8]
    sheet.cell(row=new_row, column=10).value = statistics_list[9]
    sheet.cell(row=new_row, column=11).value = statistics_list[10]
    sheet.cell(row=new_row, column=12).value = statistics_list[11]
    statistics_table.save('statistics.xlsx')
    return "Данные успешно сохранены"


def get_table_data(sheet, row):
    symbol = sheet.cell(row=row, column=1).value
    coin_link = sheet.cell(row=row, column=2).value
    order_type = sheet.cell(row=row, column=3).value
    coin_price = sheet.cell(row=row, column=4).value
    price_level = sheet.cell(row=row, column=5).value
    days_ago = sheet.cell(row=row, column=6).value
    level_touches = sheet.cell(row=row, column=7).value
    level_volume = sheet.cell(row=row, column=8).value
    average_buy_volume = sheet.cell(row=row, column=9).value
    average_sell_volume = sheet.cell(row=row, column=10).value
    average_total_volume = sheet.cell(row=row, column=11).value
    info_date = int(sheet.cell(row=row, column=12).value.timestamp() * 1000)
    coin_info = [symbol, coin_link, order_type, coin_price, price_level, days_ago, level_touches, level_volume,
                 average_buy_volume, average_sell_volume, average_total_volume, info_date]
    return coin_info


def create_result_text(analysis_result):
    three_percent_result, five_percent_result, two_percent_result = analysis_result
    result_text = [0, 0, 0]
    if three_percent_result != 0 and three_percent_result < two_percent_result != 0:
        result_text[0] = three_percent_result
    if five_percent_result != 0 and five_percent_result < two_percent_result != 0:
        result_text[1] = five_percent_result
    if two_percent_result != 0 and three_percent_result == 0 and five_percent_result == 0:
        result_text[2] = two_percent_result
    if three_percent_result != 0 and two_percent_result == 0:
        result_text[0] = three_percent_result
    if five_percent_result != 0 and two_percent_result == 0:
        result_text[1] = five_percent_result
    if three_percent_result != 0 and three_percent_result > two_percent_result != 0:
        result_text[2] = two_percent_result
    if five_percent_result != 0 and three_percent_result > two_percent_result  != 0 and five_percent_result > two_percent_result != 0:
        result_text[2] = two_percent_result
    return result_text


def get_candles(symbol, start_time):
    end_time = int(time.time() * 1000)
    url = f"https://fapi.binance.com/fapi/v1/klines"
    params = {
        'symbol': symbol,
        'interval': "1m",
        'startTime': start_time,
        'endTime': end_time,
        'limit': 241
    }
    res = get_data(url, params)
    return res


def analysis_candles(candles, coin_price, price_level, order_type):
    three_percent_result = 0
    five_percent_result = 0
    two_percent_result = 0
    if order_type == 'asks':
        three_percent_profit = price_level * 1.03
        five_percent_profit = price_level * 1.05
        two_percent_loss = price_level * 0.97
        for candle in candles:
            if three_percent_profit < float(candle[2]):
                minutes = candles.index(candle)
                if three_percent_result == 0:
                    three_percent_result = minutes + 1
            if five_percent_profit < float(candle[2]):
                minutes = candles.index(candle)
                if five_percent_result == 0:
                    five_percent_result = minutes + 1
            if two_percent_loss > float(candle[3]):
                minutes = candles.index(candle)
                if two_percent_result == 0:
                    two_percent_result = minutes + 1
    elif order_type == 'bids':
        three_percent_profit = price_level * 0.97
        five_percent_profit = price_level * 0.95
        two_percent_loss = price_level * 1.03
        for candle in candles:
            if three_percent_profit > float(candle[3]):
                minutes = candles.index(candle)
                if three_percent_result == 0:
                    three_percent_result = minutes + 1
            if five_percent_profit > float(candle[3]):
                minutes = candles.index(candle)
                if five_percent_result == 0:
                    five_percent_result = minutes + 1
            if two_percent_loss < float(candle[2]):
                minutes = candles.index(candle)
                if two_percent_result == 0:
                    two_percent_result = minutes + 1
    analysis_result = [three_percent_result, five_percent_result, two_percent_result]
    return analysis_result


def get_volume(symbol, start_time):
    # print(symbol)
    url = f"https://fapi.binance.com/futures/data/takerlongshortRatio"
    params = {
        'symbol': symbol,
        'period': "5m",
        'limit': 400,
        'startTime': start_time - 15300000,
        'endTime': start_time + 15300000
    }
    res = get_data(url, params)
    return res


def analysis_volumes(volumes, info_date, result_text, order_type):
    signal_time = False

    average_first_ten_buy_volume = 0
    average_first_ten_sell_volume = 0
    average_first_ten_sum_volume = 0

    before_signal_buy_volume = 0
    before_signal_sell_volume = 0
    before_signal_volume_counter = 0

    during_signal_buy_volume = 0
    during_signal_sell_volume = 0
    during_signal_volume_counter = 0

    level_signal_buy_volume = 0
    level_signal_sell_volume = 0
    level_signal_sum_volume = 0
    for volume in volumes:
        if volume['timestamp'] < info_date:
            before_signal_buy_volume += float(volume['buyVol'])
            before_signal_sell_volume += float(volume['sellVol'])
            before_signal_volume_counter += 1
            if before_signal_volume_counter == 10:
                try:
                    average_first_ten_buy_volume = before_signal_buy_volume / before_signal_volume_counter
                    average_first_ten_sell_volume = before_signal_sell_volume / before_signal_volume_counter
                except ZeroDivisionError:
                    pass
                average_first_ten_sum_volume = average_first_ten_buy_volume + average_first_ten_sell_volume
            if signal_time:
                during_signal_buy_volume += float(volume['buyVol'])
                during_signal_sell_volume += float(volume['sellVol'])
                during_signal_volume_counter += 1
        if volume['timestamp'] > info_date:
            if signal_time and level_signal_sum_volume == 0 and (result_text[0] != 0 or result_text[2] != 0):
                level_signal_buy_volume = float(volume['buyVol'])
                level_signal_sell_volume = float(volume['sellVol'])
                level_signal_sum_volume = level_signal_buy_volume + level_signal_sell_volume
            if not signal_time:
                if result_text[0] != 0:
                    info_date = info_date + result_text[0] * 60000
                if result_text[1] != 0:
                    info_date = info_date + result_text[1] * 60000 - result_text[0] * 60000
                if result_text[2] != 0:
                    info_date = info_date + result_text[2] * 60000
                if result_text[0] == 0 and result_text[2] == 0:
                    info_date = info_date + 245 * 60000
            if volumes[52] == volume and not signal_time:
                signal_time = volumes[52]
            elif volumes[52] != volume and not signal_time:
                signal_time = volumes[51]
    if not before_signal_volume_counter == 0:
        average_before_signal_buy_volume = before_signal_buy_volume / before_signal_volume_counter
        average_before_signal_sell_volume = before_signal_sell_volume / before_signal_volume_counter
        average_before_signal_sum_volume = average_before_signal_buy_volume + average_before_signal_sell_volume
    else:
        average_before_signal_buy_volume = 0
        average_before_signal_sell_volume = 0
        average_before_signal_sum_volume = 0

    if not during_signal_volume_counter == 0:
        average_during_signal_buy_volume = during_signal_buy_volume / during_signal_volume_counter
        average_during_signal_sell_volume = during_signal_sell_volume / during_signal_volume_counter
        average_during_signal_sum_volume = average_during_signal_buy_volume + average_during_signal_sell_volume
    else:
        average_during_signal_buy_volume = 0
        average_during_signal_sell_volume = 0
        average_during_signal_sum_volume = 0
    if order_type == 'asks':
        volumes_list = [average_first_ten_buy_volume, average_first_ten_sum_volume,
                        average_before_signal_buy_volume, average_before_signal_sum_volume,
                        average_during_signal_buy_volume, average_during_signal_sum_volume,
                        level_signal_buy_volume, level_signal_sum_volume]
        return volumes_list
    if order_type == 'bids':
        volumes_list = [average_first_ten_sell_volume, average_first_ten_sum_volume,
                        average_before_signal_sell_volume, average_before_signal_sum_volume,
                        average_during_signal_sell_volume, average_during_signal_sum_volume,
                        level_signal_sell_volume, level_signal_sum_volume]
        return volumes_list


def get_candles_for_check(symbol, info_date, interval, time_delta):
    url = f"https://api.binance.com/api/v3/klines"
    params = {
        'symbol': symbol,
        'interval': interval,
        'startTime': info_date - time_delta,
        'endTime': info_date,
        'limit': 200
    }
    res = get_data(url, params)
    return res


def check_day_prices(symbol, info_date, price_level, order_type):
    h_candles = get_candles_for_check(symbol, info_date - 3600000, "2h", 172800000)
    m_candles = get_candles_for_check(symbol, info_date, "1m", 7200000)
    if order_type == 'asks':
        for price in h_candles:
            if float(price[2]) > price_level:
                return False
        for price in m_candles:
            if float(price[2]) > price_level:
                return False
    if order_type == 'bids':
        for price in h_candles:
            if float(price[3]) < price_level:
                return False
        for price in m_candles:
            if float(price[3]) < price_level:
                return False
    return True


def collect_stats():
    statistics_table = openpyxl.load_workbook(filename="statistics_test.xlsx")
    try:
        sheet = statistics_table['info']
        row_count = sheet.max_row + 1
        for row in range(2, row_count):
            if not sheet.cell(row=row, column=15).value:
                symbol, coin_link, order_type, coin_price, price_level, days_ago, level_touches, level_volume, \
                average_buy_volume, average_sell_volume, average_total_volume, info_date = get_table_data(sheet, row)
                candles = get_candles(symbol, info_date)
                analysis_candles_result = analysis_candles(candles, coin_price, price_level, order_type)
                result_text = create_result_text(analysis_candles_result)
                volumes = get_volume(symbol, info_date)
                analysis_volumes_result = analysis_volumes(volumes, info_date, result_text, order_type)
                day_prices = check_day_prices(symbol, info_date, price_level, order_type)
                if day_prices:
                    sheet.cell(row=row, column=26).value = "Подходит"
                else:
                    sheet.cell(row=row, column=26).value = "Не подходит"
                sheet.cell(row=row, column=13).value = analysis_volumes_result[0]
                sheet.cell(row=row, column=14).value = analysis_volumes_result[1]
                sheet.cell(row=row, column=15).value = analysis_volumes_result[2]
                sheet.cell(row=row, column=16).value = analysis_volumes_result[3]
                sheet.cell(row=row, column=17).value = analysis_volumes_result[4]
                sheet.cell(row=row, column=18).value = analysis_volumes_result[5]
                sheet.cell(row=row, column=19).value = analysis_volumes_result[6]
                sheet.cell(row=row, column=20).value = analysis_volumes_result[7]
                sheet.cell(row=row, column=21).value = result_text[0]
                sheet.cell(row=row, column=22).value = result_text[1]
                sheet.cell(row=row, column=23).value = result_text[2]
    finally:
        statistics_table.save('statistics.xlsx')


if __name__ == "__main__":
    collect_stats()
